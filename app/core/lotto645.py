"""동행복권 로또6/45 당첨결과를 여러 입력 방식으로 파싱한다.

지원 형식:
1. 세로 붙여넣기(사이트에서 결과 하나씩 복사) — 회차당 11줄:
   회차("N회") / 번호6개(줄바꿈 구분) / 보너스번호 / "1등"(고정 텍스트) / 당첨자수("N명") / 1인당당첨금("N원")
2. 가로 붙여넣기(엑셀에서 표 범위를 복사 — 탭으로 구분된 한 줄에 한 회차):
   [No] 회차 번호1~6 보너스 등수 당첨자수 1인당당첨금  (No 열은 있어도 되고 없어도 됨)
3. 엑셀 파일(.xlsx) 자체 — "로또 회차별 당첨번호" 통계 다운로드 형식:
   1행 헤더(No/회차/당첨번호(6칸)/보너스/등수/당첨자수/1인당 당첨금액), 2행부터 데이터.
"""

import openpyxl


def _parse_int(value) -> int:
    """'11명', '11 명', '2,441,919,375 원'처럼 숫자 외 문자가 붙은 문자열, 또는 엑셀 셀의 숫자값(int/float)을 정수로 변환.
    문자열은 숫자만 뽑아내는 방식이라, 숫자 타입은 먼저 반올림 변환해서 소수점이 문자열 파싱에 섞여 들어가지 않게 한다
    (예: 1236.0을 str()로 바꾸면 '1236.0' → 숫자만 뽑으면 '12360'이 되는 버그 방지).
    """
    if isinstance(value, bool):
        raise ValueError(f"숫자가 아님: {value}")
    if isinstance(value, (int, float)):
        return int(round(value))
    digits = ''.join(ch for ch in str(value) if ch.isdigit())
    if not digits:
        raise ValueError(f"숫자를 찾을 수 없음: {value}")
    return int(digits)


def _build_round(round_raw, nums_raw, bonus_raw, winners_raw, prize_raw) -> dict:
    """공통 필드 파싱. nums_raw: 길이 6인 문자열/숫자 리스트."""
    round_no = str(_parse_int(round_raw))
    nums = [_parse_int(n) for n in nums_raw]
    if len(set(nums)) != 6:
        raise ValueError(f"번호 6개 파싱 실패(중복 또는 개수 오류): {nums}")
    bonus = _parse_int(bonus_raw)
    winners = _parse_int(winners_raw)
    prize = _parse_int(prize_raw)
    return {'round': round_no, 'nums': nums, 'bonus': bonus, 'winners': winners, 'prize': prize}


def parse_lotto645_block(text: str) -> tuple:
    """붙여넣은 텍스트를 파싱해 회차 리스트로 변환 (세로 형식과 가로/탭 형식을 자동 판별).
    반환: (rounds, errors) — rounds는 save_lotto645_rounds()에 바로 넘길 수 있는 dict 리스트.
    """
    raw_lines = [l for l in text.splitlines() if l.strip()]
    if not raw_lines:
        return [], []

    # 탭이 포함된 줄이 있으면 "한 줄 = 한 회차"(엑셀에서 표 범위를 복사한 형태)로 간주
    if any('\t' in l for l in raw_lines):
        return _parse_flat_lines(raw_lines)
    return _parse_vertical_lines([l.strip() for l in raw_lines])


def _parse_flat_lines(lines: list) -> tuple:
    """탭으로 구분된 한 줄짜리 회차 여러 개. [No] 회차 번호1~6 보너스 등수 당첨자수 1인당당첨금"""
    rounds = []
    errors = []
    for row_no, line in enumerate(lines, start=1):
        fields = [f.strip() for f in line.split('\t') if f.strip() != '']
        try:
            if len(fields) == 12:
                fields = fields[1:]  # 맨 앞 No(순번) 열 버림
            if len(fields) != 11:
                raise ValueError(f"열 개수가 11개(또는 No 포함 12개)가 아님: {len(fields)}개")
            round_raw, n1, n2, n3, n4, n5, n6, bonus_raw, _grade, winners_raw, prize_raw = fields
            rounds.append(_build_round(round_raw, (n1, n2, n3, n4, n5, n6), bonus_raw, winners_raw, prize_raw))
        except (ValueError, IndexError) as e:
            errors.append(f"{row_no}번째 줄 오류 — {e}")
    return rounds, errors


def _parse_vertical_lines(lines: list) -> tuple:
    """줄바꿈으로 구분된 11줄 단위 회차 여러 개(사이트에서 결과 하나씩 복사한 형태)."""
    rounds = []
    errors = []
    for i in range(0, len(lines), 11):
        chunk = lines[i:i + 11]
        if len(chunk) < 11:
            errors.append(f"마지막 {len(chunk)}줄은 11줄 단위를 채우지 못해 건너뜀: {' / '.join(chunk)}")
            break

        round_raw, n1, n2, n3, n4, n5, n6, bonus_raw, _grade, winners_raw, prize_raw = chunk
        block_no = i // 11 + 1
        try:
            rounds.append(_build_round(round_raw.rstrip('회'), (n1, n2, n3, n4, n5, n6), bonus_raw, winners_raw, prize_raw))
        except (ValueError, IndexError) as e:
            errors.append(f"{block_no}번째 블록 오류 — {e}")

    return rounds, errors


def parse_lotto645_excel_rows(rows) -> tuple:
    """엑셀에서 읽은 행(값 튜플) 이터러블을 파싱. 1행은 헤더로 간주하고 건너뜀.
    열 순서(동행복권 "로또 회차별 당첨번호" 통계 다운로드 기준):
    No, 회차, 번호1~6, 보너스, 등수, 당첨자수, 1인당당첨금
    반환: (rounds, errors)
    """
    rounds = []
    errors = []
    data_rows = list(rows)[1:]  # 헤더 행 제외
    for excel_row_no, row in enumerate(data_rows, start=2):  # 실제 엑셀 행 번호(헤더가 1행)
        if row is None or all(c is None for c in row):
            continue
        try:
            cells = list(row)
            if len(cells) < 12:
                raise ValueError(f"열 개수 부족(12개 필요): {len(cells)}개")
            rounds.append(_build_round(cells[1], cells[2:8], cells[8], cells[10], cells[11]))
        except (ValueError, IndexError, TypeError) as e:
            errors.append(f"{excel_row_no}행 오류 — {e}")
    return rounds, errors


def parse_lotto645_excel(file_obj) -> tuple:
    """업로드된 .xlsx 파일 객체(파일 경로 또는 파일 스트림)를 파싱.
    반환: (rounds, errors)
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    return parse_lotto645_excel_rows(ws.iter_rows(values_only=True))
