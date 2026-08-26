"""동행복권 연금복권720+ 당첨결과를 여러 입력 방식으로 파싱한다.

당첨번호 구성: 조(1~5) + 6자리 번호(000000~999999, 앞자리 0 보존 필요 — 문자열로 다룸).

지원 형식:
1. 세로 붙여넣기(사이트에서 결과 하나씩 복사) — 회차당 9줄:
   회차("N회") / 조 / "조"(고정 텍스트) / 번호 6자리(한 줄에 숫자 하나씩, 줄바꿈 구분)
2. 가로 붙여넣기(엑셀에서 표 범위를 복사 — 탭으로 구분된 한 줄에 한 회차):
   [No] 회차 조 당첨번호(6자리)  (No 열은 있어도 되고 없어도 됨)
3. 엑셀 파일(.xlsx) 자체 — "연금복권720+ 회차별 당첨번호" 통계 다운로드 형식:
   1행 헤더(No/회차/조/당첨번호), 2행부터 데이터.
"""

import openpyxl


def _parse_int(value) -> int:
    """숫자 외 문자가 붙은 문자열, 또는 엑셀 셀의 숫자값(int/float)을 정수로 변환."""
    if isinstance(value, bool):
        raise ValueError(f"숫자가 아님: {value}")
    if isinstance(value, (int, float)):
        return int(round(value))
    digits = ''.join(ch for ch in str(value) if ch.isdigit())
    if not digits:
        raise ValueError(f"숫자를 찾을 수 없음: {value}")
    return int(digits)


def _parse_number(value) -> str:
    """당첨번호를 6자리 문자열로 변환(앞자리 0 보존을 위해 zfill). 숫자 6자리를 벗어나면 오류."""
    if isinstance(value, bool):
        raise ValueError(f"번호가 아님: {value}")
    if isinstance(value, float):
        value = int(round(value))
    digits = ''.join(ch for ch in str(value) if ch.isdigit())
    if not digits or len(digits) > 6:
        raise ValueError(f"번호는 숫자 6자리여야 함: {value}")
    return digits.zfill(6)


def _build_round(round_raw, group_raw, number_raw) -> dict:
    """공통 필드 파싱."""
    round_no = str(_parse_int(round_raw))
    group = _parse_int(group_raw)
    if not (1 <= group <= 5):
        raise ValueError(f"조는 1~5 사이여야 함: {group}")
    number = _parse_number(number_raw)
    return {'round': round_no, 'group': group, 'number': number}


def parse_pension720_block(text: str) -> tuple:
    """붙여넣은 텍스트를 파싱해 회차 리스트로 변환 (세로 형식과 가로/탭 형식을 자동 판별).
    반환: (rounds, errors) — rounds는 save_pension720_rounds()에 바로 넘길 수 있는 dict 리스트.
    """
    raw_lines = [l for l in text.splitlines() if l.strip()]
    if not raw_lines:
        return [], []

    # 탭이 포함된 줄이 있으면 "한 줄 = 한 회차"(엑셀에서 표 범위를 복사한 형태)로 간주
    if any('\t' in l for l in raw_lines):
        return _parse_flat_lines(raw_lines)
    return _parse_vertical_lines([l.strip() for l in raw_lines])


def _parse_flat_lines(lines: list) -> tuple:
    """탭으로 구분된 한 줄짜리 회차 여러 개. [No] 회차 조 당첨번호(6자리, 또는 자리별로 6칸 분리)"""
    rounds = []
    errors = []
    for row_no, line in enumerate(lines, start=1):
        fields = [f.strip() for f in line.split('\t') if f.strip() != '']
        try:
            # No(순번) 열이 있으면 버림 — 회차/조/당첨번호 3열 기준으로 앞에 남는 열 제거
            if len(fields) == 9:
                fields = fields[1:]  # No, 회차, 조, 숫자6칸 → 회차, 조, 숫자6칸
            if len(fields) == 8:
                round_raw, group_raw, *digits = fields
                number_raw = ''.join(digits)
            elif len(fields) == 4:
                fields = fields[1:]  # No, 회차, 조, 당첨번호 → 회차, 조, 당첨번호
                round_raw, group_raw, number_raw = fields
            elif len(fields) == 3:
                round_raw, group_raw, number_raw = fields
            else:
                raise ValueError(f"열 개수가 맞지 않음(3, 4, 8, 9개 지원): {len(fields)}개")
            rounds.append(_build_round(round_raw, group_raw, number_raw))
        except (ValueError, IndexError) as e:
            errors.append(f"{row_no}번째 줄 오류 — {e}")
    return rounds, errors


def _parse_vertical_lines(lines: list) -> tuple:
    """줄바꿈으로 구분된 9줄 단위 회차 여러 개(사이트에서 결과 하나씩 복사한 형태).
    회차 / 조 / "조"(고정 텍스트, 무시) / 번호 6자리(한 줄에 하나씩)
    """
    rounds = []
    errors = []
    for i in range(0, len(lines), 9):
        chunk = lines[i:i + 9]
        if len(chunk) < 9:
            errors.append(f"마지막 {len(chunk)}줄은 9줄 단위를 채우지 못해 건너뜀: {' / '.join(chunk)}")
            break

        round_raw, group_raw, _jo_literal, *digits = chunk
        block_no = i // 9 + 1
        try:
            number_raw = ''.join(digits)
            rounds.append(_build_round(round_raw.rstrip('회').strip(), group_raw, number_raw))
        except (ValueError, IndexError) as e:
            errors.append(f"{block_no}번째 블록 오류 — {e}")

    return rounds, errors


def parse_pension720_excel_rows(rows) -> tuple:
    """엑셀에서 읽은 행(값 튜플) 이터러블을 파싱. 1행은 헤더로 간주하고 건너뜀.
    열 순서(동행복권 "연금복권720+ 회차별 당첨번호" 통계 다운로드 기준): No, 회차, 조, 당첨번호
    반환: (rounds, errors)
    """
    rounds = []
    errors = []
    data_rows = list(rows)[1:]  # 헤더 행 제외
    for excel_row_no, row in enumerate(data_rows, start=2):  # 실제 엑셀 행 번호(헤더가 1행)
        if row is None or all(c is None for c in row):
            continue
        try:
            cells = list(row)
            if len(cells) < 4:
                raise ValueError(f"열 개수 부족(4개 필요): {len(cells)}개")
            rounds.append(_build_round(cells[1], cells[2], cells[3]))
        except (ValueError, IndexError, TypeError) as e:
            errors.append(f"{excel_row_no}행 오류 — {e}")
    return rounds, errors


def parse_pension720_excel(file_obj) -> tuple:
    """업로드된 .xlsx 파일 객체(파일 경로 또는 파일 스트림)를 파싱.
    반환: (rounds, errors)
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    return parse_pension720_excel_rows(ws.iter_rows(values_only=True))
